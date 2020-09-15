#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date      : 2020/7/16 22:03
# Author    : zzh
# Python :v 3.7.3

import openpyxl

"""
读取Excel文件的方式：
            列表嵌套列表
"""


class ReadExcelDataList:
    def get_excel_data(self):
        # 按照指定路径加载excel文件并打开
        wd = openpyxl.load_workbook('../TestData/userlogindata.xlsx')
        # print(wd)
        # 指定具体的sheet页
        ws = wd['测试用例']
        all_cases = []
        # 得到Excel文件中测数据范围
        data_range = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
        for row in data_range:
            # 按行读取单元格
            # print(row)
            # 通过列表解析获取单元格的值生成列表
            row_list = [cell.value for cell in row]
            all_cases.append(row_list)
        return all_cases


if __name__ == '__main__':
    rd = ReadExcelDataList().get_excel_data()
    print(rd)
