#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date      : 2020/7/19 22:22
# Author    : zzh
# Python :v 3.7.3
import openpyxl

"""
读取Excel文件方式：
        列表嵌套字典
        wb-工作簿
        成功excel文件 然后顺利读取xlsx文件
        wb = openpyxl.load_workbook('../data/data.xlsx')
        加r,\\,/ ,以下路径表示方法是绝对路径,缺点：灵活性差  框架：架构一样的 ，相同的
        绝对路径：     base_dir                        +/data/ + 'data.xlsx'
        os.sep 自动去匹配你的系统的/,\(自动区分mac还是windows)
         D:/company/auto/20200606/lesson6/lesson6_2/=》如何获取这个路径？？？？
        base_dir = os.path.dirname(os.path.dirname(__file__))
        wb = openpyxl.load_workbook(base_dir+'%sdata%s' % (os.sep,os.sep)+'data.xlsx')

"""


class ReadExcelDataDict:
    def get_excel_data(self, head):
        """获取Excel文件中的数据"""
        # 根据指定路径加载并打开Excel文件
        wd = openpyxl.load_workbook('../TestData/userlogindata.xlsx')
        ws = wd['测试用例']
        all_cases = []

        for row in range(2, ws.max_row + 1):
            """
            首行为列名，故数据从第二行开始，到最大行加1
            """
            # 控制列数
            col = 1
            dt = {}
            for key in head:
                """head里面有多少元素，代表有几列"""
                dt[key] = ws.cell(row, col).value
                col += 1
            all_cases.append(dt)
        return all_cases


if __name__ == '__main__':
    all_cases = ReadExcelDataDict().get_excel_data(['username', 'password'])
    print(all_cases)
